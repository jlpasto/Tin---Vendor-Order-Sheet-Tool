import pandas as pd
import openai
import re
import json
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os
import time
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation


# ==============================
# CONFIGURATION
# ==============================
MASTER_FILE = "Master Product.xlsx"
CLIENT_FILE = "Client Order Sheet.xlsx"
OUTPUT_FILE = "Client Order Sheet - Updated.xlsx"
SHEET_NAME_MASTER = "Sheet1"
SHEET_NAME_CLIENT = "Sheet1"

OPENAI_MODEL = "gpt-4o-mini"  # cost-efficient model
LOG_FILE = "update_log.txt"

# Load environment variables
load_dotenv()
api_key = os.getenv("OPENAI_API_KEY")
if api_key is None:
    raise ValueError("OPENAI_API_KEY is not set in .env file.")

# Initialize OpenAI client
client = openai.OpenAI(api_key=api_key)

# ==============================
# LOGGING SETUP
# ==============================
logging.basicConfig(
    filename=LOG_FILE,
    filemode="w",
    format="%(asctime)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger()


# ==============================
# AI BATCH PARSING
# ==============================

def clean_json_output(output: str) -> str:
    """
    Cleans AI output by removing ```json or ``` markers if present.
    Returns a valid JSON string.
    """
    if output.startswith("```json"):
        output = output[len("```json"):]  # remove the starting ```json
    if output.startswith("```"):
        output = output[len("```"):]  # handles case where AI uses ``` without json
    if output.endswith("```"):
        output = output[: -len("```")]  # remove closing backticks
    return output.strip()

def round_2_decimals(value):
    logging.info(f"Rounding value: {value}")
    try:
        logging.info(f"Rounding value result: {Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)}")
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        
    except (InvalidOperation, TypeError, ValueError):
        logging.info("Rounding value result: 0.00")
        return Decimal("0.00")

def normalize_text_proper(text):
    # Words to keep lowercase unless they're the first word
    small_words = {"a", "an", "and", "as", "at", "but", "by", "for", 
                   "if", "in", "nor", "of", "on", "or", "per", "so", 
                   "the", "to", "via", "vs", "yet"}
    
    words = re.split(r'(\s+)', text)  # Keeps spaces intact
    normalized_words = []

    for i, word in enumerate(words):
        if word.isspace():
            normalized_words.append(word)
        elif word.isdigit():  # Keep numbers as-is
            normalized_words.append(word)
        else:
            lower_word = word.lower()
            if i == 0 or lower_word not in small_words:  # First word or not a small word
                normalized_words.append(lower_word.capitalize())
            else:
                normalized_words.append(lower_word)

    return ''.join(normalized_words)

def format_size(unit_size, unit_uom):
    """
    Formats the unit size and unit of measure.
    - Removes unnecessary '.0' if size is a whole number.
    - Converts UOM to lowercase and trims spaces.
    """
    size_value = str(unit_size).strip() if unit_size is not None else "1"
    uom = str(unit_uom).lower().strip()

    try:
        num = float(size_value)
        # If it's a whole number, convert to int to drop ".0"
        if num.is_integer():
            size_value = str(int(num))
        else:
            size_value = str(num)
    except ValueError:
        # If size is not numeric (e.g., "N/A"), leave it as is
        pass

    return f"{size_value} {uom}"


def parse_item_descriptions_batch(items, batch_size=30, retries=2):
    """
    Batch-process item descriptions with AI.
    Each `items` element is a tuple: (UPC, Item Description).
    Returns dict {upc: {"vendor":..., "product":..., "case_pack":..., "size":...}}
    """
    all_results = {}

    for i in range(0, len(items), batch_size):
        batch = items[i:i + batch_size]
        formatted_input = "\n".join([
            f"{idx+1}. UPC: {upc} | Description: {desc}"
            for idx, (upc, desc) in enumerate(batch)
        ])

        prompt = f"""
        You are an expert at parsing product descriptions into structured data for product catalogs.

        Your task: Extract the following from the given product description:
        - upc: copy from the input exactly
        - vendor (brand/manufacturer)
        - product (unique product name without case pack and size)

        Guidelines:
        1. If any field is missing, return null for that field.
        2. The "product" field must be unique — if multiple products share the same name but differ in attributes like size or pack unit (e.g., carton vs. tray), include those attributes in the product name.
        3. Do NOT include case pack quantity or size in the product name unless needed to differentiate exactly same product name.
        4. Output ONLY valid JSON, with keys: vendor, product, case_pack, size. Exclude any text outside the JSON. (e.g. ```json)

        ### Example

        **Input:**
        Tuple of upc, descriptions:

        [(698264001101, Alderfer - Cage Free Large Brown Eggs, Pulp Carton (15 x 1 dozen))
        (698264000906, Alderfer - Cage Free Large White Eggs, Pulp Carton (15 x 1 dozen))
        (698264002207, Alderfer - Organic Large White, Pulp Carton (15 x 1 dozen))
        (698264000401, Alderfer - Organic Large Brown, Pulp Carton (15 x 1 dozen))]

        **Output:**
        [
        {{  
            "upc": 698264001101,
            "vendor": "Alderfer",
            "product": "Eggs, Large Brown, Pulp Carton - Cage Free"
        }},
        {{
            "upc": 698264000906,
            "vendor": "Alderfer",
            "product": "Eggs, Large White, Pulp Carton - Cage Free"
        }},
        {{
            "upc": 698264002207,
            "vendor": "Alderfer",
            "product": "Eggs, Large White, Pulp Carton - Organic"
        }},
        {{
            "upc": 698264000401,
            "vendor": "Alderfer",
            "product": "Eggs, Large Brown, Pulp Carton - Organic"
        }}
        ]

        ### Now process this formatted input:
        Input: "{formatted_input}"
        """

        attempt = 0
        while attempt <= retries:
            try:
                response = client.chat.completions.create(
                    model=OPENAI_MODEL,
                    messages=[
                        {"role": "system", "content": "You are a precise data parser."},
                        {"role": "user", "content": prompt}
                    ],
                )
                output = response.choices[0].message.content
                logger.info(f"Batch {i//batch_size+1} raw output: {output}")

                results = json.loads(clean_json_output(output))
                if not isinstance(results, list):
                    raise ValueError("Expected JSON array")

                for result in results:
                    upc = str(result.get("upc")).strip()
                    if upc:
                        all_results[upc] = {
                            "vendor": result.get("vendor"),
                            "product": result.get("product")
                        }
                break  # success, break retry loop
            except Exception as e:
                logger.error(f"❌ Error parsing batch {i//batch_size+1} attempt {attempt+1}: {e}")
                attempt += 1
                time.sleep(2)

    return all_results


# ==============================
# MAIN UPDATE LOGIC
# ==============================
def update_client_sheet():
    try:
        logger.info("Loading files...")
        master_df = pd.read_excel(MASTER_FILE, sheet_name=SHEET_NAME_MASTER, dtype={"UPC": str})
        client_df = pd.read_excel(CLIENT_FILE, sheet_name=SHEET_NAME_CLIENT, dtype={"UPC": str})

        # Build list of (UPC, Description) to parse
        items_to_parse = [
            (str(row["UPC"]).strip(), row["Item Description"])
            for _, row in master_df.iterrows()
            if pd.notna(row.get("UPC")) and pd.notna(row.get("Item Description"))
        ]

        logger.info(f"Parsing {len(items_to_parse)} items with AI in batches...")
        parsed_results = parse_item_descriptions_batch(items_to_parse)

        logger.info("Building master product dictionary...")
        master_products = {}
        for _, row in master_df.iterrows():
            upc = str(row.get("UPC")).strip() if pd.notna(row.get("UPC")) else None
            if not upc:
                continue
            parsed = parsed_results.get(upc, {})
            master_products[upc] = {
                "vendor": normalize_text_proper(row.get("Manufacturer Name")) if pd.notna(row.get("Manufacturer Name")) else "",
                "product": re.sub(r"\bwith\b", "-", parsed.get("product"), flags=re.IGNORECASE),
                "case_pack": row.get("Pack Size") if pd.notna(row.get("Pack Size")) else "",
                "size": format_size(row.get("Unit Size"), row.get("Unit UOM")),
                "wholesale_case_price": row.get("Wholesale Case Price"),
                "wholesale_unit_price": row.get("Wholesale Unit Price"),
                "retail_price": row.get("Retail Unit Price (MSRP)"),
            }

        # Load workbook for editing
        wb = load_workbook(CLIENT_FILE)
        ws = wb[SHEET_NAME_CLIENT]

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        client_upcs = set()
        max_row = ws.max_row

        edited_count = 0
        added_count = 0
        deleted_count = 0

        logger.info("Processing existing rows...")
        for row_idx in range(2, max_row + 1):
            upc_value  = ws[f"F{row_idx}"].value  # Assuming UPC is column F
            upc = None
            if upc_value is None:
                upc = None
            else:
                # Convert to string and strip ".0" if float
                if isinstance(upc_value, float):
                    upc = str(int(upc_value))  # int() removes the .0 safely
                else:
                    upc = str(upc_value).strip()

            logger.info(f"Processing row: {row_idx} | UPC: {upc}")

            id_val = ws[f"A{row_idx}"].value
            if not upc:
                continue

            upc_str = str(upc).strip()
            client_upcs.add(upc_str)

            if upc_str in master_products:
                product_info = master_products[upc_str]
                edited = False

                update_map = {
                    "B": product_info["vendor"],
                    "C": product_info["product"],
                    "D": product_info["size"],
                    "E": str(int(float(product_info["case_pack"]))), # ensure case_pack is int string
                    "F": upc_str,
                    "G": f"{round_2_decimals((product_info["wholesale_case_price"])):.2f}",
                    "H": f"{round_2_decimals((product_info["wholesale_unit_price"])):.2f}",
                    "I": f"{round_2_decimals((product_info["retail_price"])):.2f}"
                }

                for col, new_val in update_map.items():
                    cell = ws[f"{col}{row_idx}"]

                    # Normalize both values to string for comparison
                    #current_val = "" if cell.value is None else str(cell.value).strip()
                    new_val_str = "" if new_val is None else str(new_val).strip()
                    current_val = f"{round(float(cell.value), 2):.2f}" if col in ("G", "H", "I") else str(cell.value).strip()
                    logging.info(f"Comparing {col}{row_idx}: current='{current_val}' vs new='{new_val_str}'")
                    if col in ("F",):  # UPC column
                        cell.value = upc_str  # always update UPC to avoid scientific notation

                    if col in ("E",):  # case_pack should be integer string
                        try:
                            new_val_str = str(int(float(new_val_str)))
                        except (ValueError, TypeError):
                            new_val_str = "0"

                    if current_val != new_val_str:
                        logger.info(f"Updated {col}{row_idx} {upc_str} from {current_val} to {new_val_str}")
                        cell.value = new_val_str
                        cell.fill = green_fill
                        edited = True

                if edited:
                    ws[f"N{row_idx}"] = "EDITED"
                    ws[f"N{row_idx}"].fill = green_fill
                    edited_count += 1
            else:
                # update the ucp to string in the sheet/ avoid scientific notation
                ws[f"F{row_idx}"].value = upc_str

                ws[f"N{row_idx}"] = "DELETE"
                for col in range(1, ws.max_column + 1):
                    ws[f"{get_column_letter(col)}{row_idx}"].fill = red_fill
                deleted_count += 1

        logger.info("Adding new products...")
        for upc, data in master_products.items():
            if upc not in client_upcs:
                max_row += 1
                ws[f"B{max_row}"] = data["vendor"]
                ws[f"C{max_row}"] = data["product"]
                ws[f"D{max_row}"] = data["size"]
                ws[f"E{max_row}"] = data["case_pack"]
                ws[f"F{max_row}"] = upc
                ws[f"G{max_row}"] = f"{round_2_decimals((data["wholesale_case_price"])):.2f}"
                ws[f"H{max_row}"] = f"{round_2_decimals((data["wholesale_unit_price"])):.2f}" 
                ws[f"I{max_row}"] = f"{round_2_decimals((data["retail_price"])):.2f}"  
                ws[f"N{max_row}"] = "ADDED"
                for col in "BCDEFGHI":
                    ws[f"{col}{max_row}"].fill = green_fill
                added_count += 1

        logger.info("Saving updated workbook...")
        wb.save(OUTPUT_FILE)
        logger.info(f"Update completed. EDITED={edited_count}, ADDED={added_count}, DELETE={deleted_count}")

    except Exception as e:
        logger.exception("Error during update process")
        raise


if __name__ == "__main__":
    update_client_sheet()
    print(f"✅ Update complete. See '{OUTPUT_FILE}' and check '{LOG_FILE}' for details.")
